from openpyxl import *


class CommonFlow:
    nodes = []
    branches = []

    def __init__(self, excel_file):
        wb = load_workbook(excel_file)
        sheet = wb.worksheets[0]
        counter = 2
        while sheet.cell(counter, 1).value is not None:
            node = CommonNode(
                ny=sheet.cell(counter, 1).value,
                name=sheet.cell(counter, 2).value,
                uhom=sheet.cell(counter, 3).value,
                nsx=sheet.cell(counter, 4).value,
                na=sheet.cell(counter, 5).value,
                pn=sheet.cell(counter, 6).value,
                qn=sheet.cell(counter, 7).value,
                pg=sheet.cell(counter, 8).value,
                qg=sheet.cell(counter, 9).value,
                vzd=sheet.cell(counter, 10).value,
                qmin=sheet.cell(counter, 11).value,
                qmax=sheet.cell(counter, 12).value,
                bsh=sheet.cell(counter, 13).value,
                options=sheet.cell(counter, 14).value
            )
            self.nodes.append(node)
            counter += 1
        sheet = wb.worksheets[1]
        counter = 2
        while sheet.cell(counter, 1).value is not None:
            branch = CommonBranch(
                ip=sheet.cell(counter, 1).value,
                iq=sheet.cell(counter, 2).value,
                np=sheet.cell(counter, 3).value,
                name=sheet.cell(counter, 4).value,
                r=sheet.cell(counter, 5).value,
                x=sheet.cell(counter, 6).value,
                b=sheet.cell(counter, 7).value,
            )
            self.branches.append(branch)
            counter += 1


class CommonBranch:
    """Ветвь, которая будет извлекаться из Растра (или Экселя)"""
    def __init__(self, r=0.0, x=0.0, b=0.0, kti=0, name="", ip=0, iq=0, np=0):
        self.r = r
        self.x = x
        self.b = b
        self.kti = kti
        self.name = name
        self.ip = ip
        self.iq = iq
        self.np = np

    def __str__(self):
        return self.name


class CommonNode:
    """Узел, который будет извлекаться из Растра (или Экселя)"""
    def __init__(self, ny=0, name="", uhom=0.0,
                 nsx=0, na=0, pn=0.0, qn=0.0,
                 pg=0.0, qg=0.0, vzd=0.0,
                 qmin=0.0, qmax=0.0, bsh=0.0, options=''):
        self.ny = ny
        self.name = name
        self.uhom = uhom
        self.nsx = nsx
        self.na = na
        self.pn = pn
        self.qn = qn
        self.pg = pg
        self.qg = qg
        self.vzd = vzd
        self.qmin = qmin
        self.qmax = qmax
        self.bsh = bsh
        self.options = options

    def __str__(self):
        return self.name
