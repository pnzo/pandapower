from model.common_elements.CommonBranch import CommonBranch
from model.common_elements.CommonNode import CommonNode
from model.common_elements.CommonLoadCharacteristic import CommonLoadCharacteristic
from openpyxl import *
from typing import List


class CommonFlow:
    nodes: List[CommonNode] = []
    branches: List[CommonBranch] = []
    load_characteristics: List[CommonLoadCharacteristic] = []

    def __init__(self, excel_file):
        wb = load_workbook(excel_file)
        sheet = wb.worksheets[0]
        counter = 2
        while sheet.cell(counter, 1).value is not None:
            node = CommonNode(
                ny=sheet.cell(counter, 1).value,
                name=str(sheet.cell(counter, 2).value),
                uhom=sheet.cell(counter, 3).value,
                nsx=sheet.cell(counter, 4).value,
                na=sheet.cell(counter, 5).value,
                pn=sheet.cell(counter, 6).value,
                qn=sheet.cell(counter, 7).value,
                pg=sheet.cell(counter, 8).value,
                qg=sheet.cell(counter, 9).value,
                vzd=sheet.cell(counter, 10).value,
                qmin=sheet.cell(counter, 11).value,
                qmax=sheet.cell(counter, 12).value,
                bsh=sheet.cell(counter, 13).value,
                options=sheet.cell(counter, 14).value
            )
            self.nodes.append(node)
            counter += 1
        sheet = wb.worksheets[1]
        counter = 2
        while sheet.cell(counter, 1).value is not None:
            branch = CommonBranch(
                ip=sheet.cell(counter, 1).value,
                iq=sheet.cell(counter, 2).value,
                np=sheet.cell(counter, 3).value,
                name=sheet.cell(counter, 4).value,
                r=sheet.cell(counter, 5).value,
                x=sheet.cell(counter, 6).value,
                b=sheet.cell(counter, 7).value,
                ktr=sheet.cell(counter, 8).value,
                kti=sheet.cell(counter, 9).value
            )
            self.branches.append(branch)
            counter += 1
        sheet = wb.worksheets[2]
        counter = 2
        while sheet.cell(counter, 1).value is not None:
            load_characteristic = CommonLoadCharacteristic(
                nsx=sheet.cell(counter, 1).value,
                p0=sheet.cell(counter, 2).value,
                p1=sheet.cell(counter, 3).value,
                p2=sheet.cell(counter, 4).value,
                q0=sheet.cell(counter, 5).value,
                q1=sheet.cell(counter, 6).value,
                q2=sheet.cell(counter, 7).value,

            )
            self.load_characteristics.append(load_characteristic)
            counter += 1
